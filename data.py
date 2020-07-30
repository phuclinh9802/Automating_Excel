import xlrd
import openpyxl
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import timeit
import xlsxwriter
import csv
from itertools import zip_longest
import random
import string
import numpy as np
import scipy
from scipy import stats
import math
import Automating_Excel.automate_hmdb as hm
import os, glob
import shutil
import time
import pandas as pd
from splinter import Browser
from selenium import webdriver
# file to be processed: Raw_data_and_steps_Diabetes_data.xlsx
# Replace 0 with empty cell
from scipy.stats import sem, t
from scipy.stats import ttest_ind, ttest_ind_from_stats
from scipy.special import stdtr





def replace_empty(lists):
    table_len = len(lists)
    element_len = len(lists[0])
    for y in range(table_len):
        for x in range(element_len):
            if lists[y][x] == 0:
                lists[y][x] = None


# create a new empty xlsx file
# def new_file():
#     workbook = xlsxwriter.Workbook('new_file.xlsx')
#     worksheet = workbook.add_worksheet()
#     for x in range()
#         worksheet.write(0,0)
#     workbook.close()

# create calculated data from original table
def new_file_calculated(lists, cols, str):
    print(str)
    if str == "Raw_data_and_steps_Diabetes_data.xlsx":
        workbook = xlsxwriter.Workbook('calculated_data_Raw.xlsx')
    elif str == "Fish Liver.xlsx":
        workbook = xlsxwriter.Workbook('calculated_data_FL.xlsx')
    elif str == "Fish Muscle.xlsx":
        workbook = xlsxwriter.Workbook('calculated_data_FM.xlsx')
    elif str == "20200625_HIV_KSHV_MZMineOutput.xlsx":
        workbook = xlsxwriter.Workbook('calculated_data_HIV.xlsx')
    worksheet = workbook.add_worksheet()

    for x in range(cols):
        worksheet.write_column(0, x, lists[x])
    workbook.close()


# Read group data only
def read_group_data(str):
    wb = xlrd.open_workbook(str)
    ws = wb.sheet_by_index(0)
    rows = ws.nrows
    cols = ws.ncols
    table = []
    count = 0

    for y in range(cols):
        record = []
        if y < cols:
            for x in range(rows):
                record.append(ws.cell(x, y).value)
            new_record = record
            table.append(new_record)

    return table




# Read group data with average: 1. Control 2. Diabetes 3. Diabetes+Insulin
def read_group_data_with_average(str):
    wb = xlrd.open_workbook(str)
    ws = wb.sheet_by_index(0)
    rows = ws.nrows
    cols = ws.ncols
    table = []
    count = 0
    average_column = calculate_average(str)

    for y in range(cols):
        record = []
        if y < cols:
            for x in range(rows):
                record.append(ws.cell(x,y).value)
            new_record = record
            table.append(new_record)

    # calculate average of group data in each row using openpyxl
    table.append(average_column)

    return table

# convert data to table (all files can be converted)
def read_all_data(file):
    wb = xlrd.open_workbook(file)
    ws = wb.sheet_by_index(0)
    rows = ws.nrows
    cols = ws.ncols
    table = []
    for y in range(cols):
        record = []
        for x in range(rows):
            record.append(ws.cell(x, y).value)
        new_record = record
        table.append(new_record)

    return table


# Read the data and calculate
def read_data(str):
    # Load excel file to calculate
    wb = xlrd.open_workbook(str)
    ws = wb.sheet_by_index(0)
    rows = ws.nrows
    cols = ws.ncols
    table = []
    count = 0

    # calculate data - Blank
    for y in range(cols):
        record = []
        count += 1
        for x in range(rows):
            if str == "Raw_data_and_steps_Diabetes_data.xlsx" or str == "Fish Muscle.xlsx":
                if 0 < y < 16:
                    if isinstance(ws.cell(x, y).value, float) and isinstance(ws.cell(x, 16).value, float):
                        if ws.cell(x, y).value - ws.cell(x, 16).value >=0:
                            record.append(ws.cell(x, y).value - ws.cell(x, 16).value)
                        else:
                            record.append(None)
                    else:
                        record.append(ws.cell(x, y).value)
                else:
                    record.append(ws.cell(x, y).value)
            elif str == "Fish Liver.xlsx":
                if 0 < y < 15:
                    if isinstance(ws.cell(x, y).value, float) and isinstance(ws.cell(x, 15).value, float):
                        if ws.cell(x, y).value - ws.cell(x, 15).value >= 0:
                            record.append(ws.cell(x, y).value - ws.cell(x, 15).value)
                        else:
                            record.append(None)
                    else:
                        record.append(ws.cell(x, y).value)
                else:
                    record.append(ws.cell(x, y).value)
            elif str == "20200625_HIV_KSHV_MZMineOutput.xlsx":
                if 1 < y < 53:
                    if isinstance(ws.cell(x, y).value, float) and isinstance(ws.cell(x, 1).value, float):
                        if ws.cell(x, y).value - ws.cell(x, 1).value >= 0:
                            record.append(ws.cell(x, y).value - ws.cell(x, 1).value)
                        else:
                            record.append(None)
                    else:
                        record.append(ws.cell(x, y).value)
                else:
                    record.append(ws.cell(x, y).value)
        new_record = record
        table.append(new_record)

    # replace 0 with empty cell
    replace_empty(table)

    # separating calculations to another xlsx file
    new_file_calculated(table, count, str)

    return table


# produce table only with new count (not generating new xlsx file)
def produce_table_only(string):
    original_table = read_data("Raw_data_and_steps_Diabetes_data.xlsx")
    separated_table = separating_group(original_table, string)
    return


# produce a new data with count
def produce_count_data(string, filename):
    original_table = read_data(filename)
    separated_table = separating_group(original_table, string, filename)
    if filename == "Raw_data_and_steps_Diabetes_data.xlsx":
        if string == "Control":
            workbook = xlsxwriter.Workbook('Control_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "Diabetes":
            workbook = xlsxwriter.Workbook('Diabetes_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "Diabetes+Insulin":
            workbook = xlsxwriter.Workbook('Diabetes_Insulin_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        else:
            print("Please Try Again!")
    elif filename == "Fish Liver.xlsx":
        if string == "COF":
            workbook = xlsxwriter.Workbook('COF_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "PPF":
            workbook = xlsxwriter.Workbook('PPF_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "TAMF":
            workbook = xlsxwriter.Workbook('TAMF_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        else:
            print("Please Try Again!")
    elif filename == "Fish Muscle.xlsx":
        if string == "COF":
            workbook = xlsxwriter.Workbook('COF_Muscle_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "PP":
            workbook = xlsxwriter.Workbook('PP_Muscle_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "TAMM":
            workbook = xlsxwriter.Workbook('TAMM_Muscle_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
    elif filename == "20200625_HIV_KSHV_MZMineOutput.xlsx":
        if string == "QC":
            workbook = xlsxwriter.Workbook('QC_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "Before":
            workbook = xlsxwriter.Workbook('Before_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "After":
            workbook = xlsxwriter.Workbook('After_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "KSHV+HIV+":
            workbook = xlsxwriter.Workbook('KSHV+HIV+_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "KSHV+HIV-":
            workbook = xlsxwriter.Workbook('KSHV+HIV-_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        elif string == "KSHV-HIV-":
            workbook = xlsxwriter.Workbook('KSHV-HIV-_Group.xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(separated_table)):
                worksheet.write_column(0, x, separated_table[x])
            workbook.close()
        else:
            print("Please Try Again!")



# Build a GUI to automatically calculate and generate a new separated file
def tkinter_window():

    window = Tk()
    # frame = Frame(window)
    window.title("Calculating Metabolomic Data")

    window.geometry('900x450')

    tab_control = ttk.Notebook(window)
    tab_control.pack(expand=YES, fill="both")
    tab1 = ttk.Frame(tab_control)
    tab_control.add(tab1, text="Calculate Step 1")


    tab2 = ttk.Frame(tab_control)
    tab_control.add(tab2, text="Separate Group")

    tab3 = ttk.Frame(tab_control)
    tab_control.add(tab3, text="Check Percentage")

    tab4 = ttk.Frame(tab_control)
    tab_control.add(tab4, text="Final Data")

    tab5 = ttk.Frame(tab_control)
    tab_control.add(tab5, text="Get p value")

    tab6 = ttk.Frame(tab_control)
    tab_control.add(tab6, text="Up/Down-regulated")

    tab7 = ttk.Frame(tab_control)
    tab_control.add(tab7, text="Automate HMDB")


    # tab 1
    lbl = Label(tab1, text="Choose file to calculate")
    lbl.pack(padx=2, pady=2)

    rad = IntVar()
    radio_1 = Radiobutton(tab1, text="Raw_data_and_steps_Diabetes_data.xlsx", variable=rad, value=0)
    radio_1.pack(anchor=W)
    radio_2 = Radiobutton(tab1, text="Fish Liver.xlsx", variable=rad, value=1)
    radio_2.pack(anchor=W)
    radio_3 = Radiobutton(tab1, text="Fish Muscle.xlsx", variable=rad, value=2)
    radio_3.pack(anchor=W)
    radio_4 = Radiobutton(tab1, text="20200625_HIV_KSHV_MZMineOutput.xlsx", variable=rad, value=3)
    radio_4.pack(anchor=W)

    # generate a new xlsx file
    def clicked():
        res = "File has been entered."
        if rad.get() == 0:
            read_data("Raw_data_and_steps_Diabetes_data.xlsx")
        elif rad.get() == 1:
            read_data("Fish Liver.xlsx")
        elif rad.get() == 2:
            read_data("Fish Muscle.xlsx")
        elif rad.get() == 3:
            read_data("20200625_HIV_KSHV_MZMineOutput.xlsx")
        messagebox.showinfo('Success!', res)

    btn = Button(tab1, text="Generate", command=clicked)
    btn.pack(padx=5, pady=5)

    # tab 2
    lbl_2 = Label(tab2, text="Group Name: Control, Diabetes, Diabetes+Insulin, COF, PPF, TAMF, PP, TAMM")
    lbl_2.pack(padx=2, pady=2)

    txt_2 = Entry(tab2, width=40)
    txt_2.pack(padx=2, pady=2)

    txt_2_1 = Entry(tab2, width=40)
    txt_2_1.pack(padx=2, pady=2)

    def separate():
        res = "Group name has been entered"
        failed_msg = 'There is no such group. Please try again'
        text = txt_2.get()
        text_1 = txt_2_1.get()
        # if text != 'Control' or text != 'Diabetes' or text != 'Diabetes+Insulin':
        #     messagebox.showinfo('Failed!', failed_msg)
        if text == "COF":
            if text_1 == "Fish Liver.xlsx":
                produce_count_data(text, "Fish Liver.xlsx")
            elif text_1 == "Fish Muscle.xlsx":
                produce_count_data(text, "Fish Muscle.xlsx")
        else:
            produce_count_data(text, text_1)
            messagebox.showinfo('Success!', res)

    btn_2 = Button(tab2, text="Generate", command=separate)
    btn_2.pack(padx=5, pady=5)
    #
    # tab 3 - check percentage
    lbl_3 = Label(tab3, text="Group Name: Control, Diabetes, Diabetes+Insulin")
    lbl_3.pack(padx=2, pady=2)

    txt_3 = Entry(tab3, width=40)
    txt_3.pack(padx=2,pady=2)

    def check():
        text = txt_3.get()
        res = "Perfect! The file is being processed."
        failed = "Either the group does not exist or the file have not been created. Please try again."
        if text == "Control":
            check_percentage("Control_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "Diabetes":
            check_percentage("Diabetes_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "Diabetes+Insulin":
            check_percentage("Diabetes_Insulin_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "COF":
            check_percentage("COF_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "PPF":
            check_percentage("PPF_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "TAMF":
            check_percentage("TAMF_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "COF Muscle":
            check_percentage("COF_Muscle_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "PP":
            check_percentage("PP_Muscle_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "TAMM":
            check_percentage("TAMM_Muscle_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "QC":
            check_percentage("QC_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "Before":
            check_percentage("Before_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "After":
            check_percentage("After_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "KSHV+HIV+":
            check_percentage("KSHV+HIV+_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "KSHV+HIV-":
            check_percentage("KSHV+HIV-_Group.xlsx")
            messagebox.showinfo('Success!', res)
        elif text == "KSHV-HIV-":
            check_percentage("KSHV-HIV-_Group.xlsx")
            messagebox.showinfo('Success!', res)
        else:
            messagebox.showinfo('Failed!', failed)

    btn_3 = Button(tab3, text="Generate", command=check)
    btn_3.pack(padx=5, pady=5)

    # tab 4 - save to csv
    lbl_4 = Label(tab4, text="Generate Final Data")
    lbl_4.pack(padx=2, pady=2)

    txt_4 = Entry(tab4, width=40)
    txt_4.pack(padx=2, pady=2)


    def final():
        start = timeit.default_timer()
        res = "Perfect! The file is being processed."
        final_table = []
        # produce data after checking percentage in tables
        control_table = read_group_data_with_average("Control_Group.xlsx")
        diabetes_table = read_group_data_with_average("Diabetes_Group.xlsx")
        diabetes_insulin_table = read_group_data_with_average("Diabetes_Insulin_Group.xlsx")
        cof_table = read_group_data_with_average("COF_Group.xlsx")
        ppf_table = read_group_data_with_average("PPF_Group.xlsx")
        tamf_table = read_group_data_with_average("TAMF_Group.xlsx")
        cof_muscle = read_group_data_with_average("COF_Muscle_Group.xlsx")
        pp_table = read_group_data_with_average("PP_Muscle_Group.xlsx")
        tamm_table = read_group_data_with_average("TAMM_Muscle_Group.xlsx")
        qc_table = read_group_data_with_average("QC_Group.xlsx")
        before_table = read_group_data_with_average("Before_Group.xlsx")
        after_table = read_group_data_with_average("After_Group.xlsx")
        kshv_plus_hiv_plus = read_group_data_with_average("KSHV+HIV+_Group.xlsx")
        kshv_plus_hiv_minus = read_group_data_with_average("KSHV+HIV-_Group.xlsx")
        kshv_minus_hiv_minus = read_group_data_with_average("KSHV-HIV-_Group.xlsx")
        file = txt_4.get()
        # remove rows that have no data
        i = 1
        if file == "Raw":
            while i < len(control_table[0]):
                if control_table[len(control_table) - 1][i] == 1 and diabetes_table[len(diabetes_table) - 1][i] == 1 and diabetes_insulin_table[len(diabetes_insulin_table) - 1][i] == 1:
                    remove_rows(control_table, i)
                    remove_rows(diabetes_table, i)
                    remove_rows(diabetes_insulin_table, i)
                else:
                    i += 1

            # append to a big table
            for x in range(len(control_table)):
                final_table.append(control_table[x])
            for x in range(1, len(diabetes_table)):
                final_table.append(diabetes_table[x])
            for x in range(1, len(diabetes_insulin_table)):
                final_table.append(diabetes_insulin_table[x])
            save_csv(final_table, file)
        elif file == "FL":
            j = 1
            k = 1
            m = 1
            while j < len(cof_table[0]):
                if cof_table[len(cof_table) - 1][j] == 1 and ppf_table[len(ppf_table) - 1][j] == 1 and tamf_table[len(tamf_table) - 1][j] == 1:
                    remove_rows(cof_table, j)
                    remove_rows(ppf_table, j)
                    remove_rows(tamf_table, j)
                else:
                    j += 1

            # while k < len(ppf_table):
            #     if ppf_table[len(ppf_table) - 1][k] == 1:
            #         remove_rows(ppf_table, k)
            #     else:
            #         k += 1
            # while m < len(tamf_table):
            #     if tamf_table[len(tamf_table) - 1][k] == 1:
            #         remove_rows(tamf_table, m)
            #     else:
            #         m += 1
            # append to a big table
            for x in range(len(cof_table)):
                final_table.append(cof_table[x])
            for x in range(1, len(ppf_table)):
                final_table.append(ppf_table[x])
            for x in range(1, len(tamf_table)):
                final_table.append(tamf_table[x])
            save_csv(final_table, file)
        elif file == "FM":
            j = 1

            while j < len(cof_muscle[0]):
                if cof_muscle[len(cof_muscle) - 1][j] == 1 and pp_table[len(pp_table) - 1][j] == 1 and tamm_table[len(tamm_table) - 1][j] == 1:
                    remove_rows(cof_muscle, j)
                    remove_rows(pp_table, j)
                    remove_rows(tamm_table, j)
                else:
                    j += 1

            # append to a big table
            for x in range(len(cof_muscle)):
                final_table.append(cof_muscle[x])
            for x in range(1, len(pp_table)):
                final_table.append(pp_table[x])
            for x in range(1, len(tamm_table)):
                final_table.append(tamm_table[x])
            save_csv(final_table, file)

        elif file == "HIV":
            j = 1
            while j < len(qc_table[0]):
                if qc_table[len(qc_table) - 1][j] == 1 and before_table[len(before_table) - 1][j] == 1 and after_table[len(after_table) - 1][j] == 1 and kshv_plus_hiv_plus[len(kshv_plus_hiv_plus) - 1][j] == 1 and kshv_plus_hiv_minus[len(kshv_plus_hiv_minus) - 1][j] == 1 and kshv_minus_hiv_minus[len(kshv_minus_hiv_minus) - 1][j] == 1:
                    remove_rows(qc_table, j)
                    remove_rows(before_table, j)
                    remove_rows(after_table, j)
                    remove_rows(kshv_plus_hiv_plus, j)
                    remove_rows(kshv_plus_hiv_minus, j)
                    remove_rows(kshv_minus_hiv_minus, j)
                else:
                    j += 1

            # append to a big table
            for x in range(len(qc_table)):
                final_table.append(qc_table[x])
            for x in range(1, len(before_table)):
                final_table.append(before_table[x])
            for x in range(1, len(after_table)):
                final_table.append(after_table[x])
            for x in range(1, len(kshv_plus_hiv_plus)):
                final_table.append(kshv_plus_hiv_plus[x])
            for x in range(1, len(kshv_plus_hiv_minus)):
                final_table.append(kshv_plus_hiv_minus[x])
            for x in range(1, len(kshv_minus_hiv_minus)):
                final_table.append(kshv_minus_hiv_minus[x])
            save_csv(final_table, file)

        messagebox.showinfo('Success!', res)
        stop = timeit.default_timer()
        print('Time: ', stop - start)

    btn_4 = Button(tab4, text="Generate", command=final)
    btn_4.pack(padx=5, pady=5)

    # tab 5 - get p value
    lbl_5 = Label(tab5, text="Get p value - C: Control, D: Diabetes; DI: Diabetes+Insulin")
    lbl_5.pack(padx=2, pady=2)

    txt_5 = Entry(tab5, width=20)
    txt_5.pack(padx=2, pady=2)

    txt_5_2 = Entry(tab5, width=20)
    txt_5_2.pack(padx=2, pady=2)

    v = IntVar()
    radio_1 = Radiobutton(tab5, text="p value only", variable=v, value=0)
    radio_1.pack(anchor=W)
    radio_2 = Radiobutton(tab5, text="p and log2", variable=v, value=1)
    radio_2.pack(anchor=W)

    # generate file based on entry
    def pval():
        res = "File has been generated"
        failed = "Please enter the correct group"
        if v.get() == 0:
            if (txt_5.get() == "C" and txt_5_2.get() == "D") or (txt_5.get() == "D" and txt_5_2.get() == "C"):
                produce_combine_p("Control_Group.xlsx", "Diabetes_Group.xlsx", 0)
                messagebox.showinfo('Success!', res)
            elif (txt_5.get() == "C" and txt_5_2.get() == "DI") or (txt_5.get() == "DI" and txt_5_2.get() == "C"):
                produce_combine_p("Control_Group.xlsx", "Diabetes_Insulin_Group.xlsx", 0)
                messagebox.showinfo('Success!', res)
            elif (txt_5.get() == "D" and txt_5_2.get() == "DI") or (txt_5.get() == "DI" and txt_5_2.get() == "D"):
                produce_combine_p("Diabetes_Group.xlsx", "Diabetes_Insulin_Group.xlsx", 0)
                messagebox.showinfo('Success!', res)
            else:
                messagebox.showerror("Error", failed)
        if v.get() == 1:
            if (txt_5.get() == "C" and txt_5_2.get() == "D") or (txt_5.get() == "D" and txt_5_2.get() == "C"):
                produce_combine_p("Control_Group.xlsx", "Diabetes_Group.xlsx", 1)
                messagebox.showinfo('Success!', res)
            elif (txt_5.get() == "C" and txt_5_2.get() == "DI") or (txt_5.get() == "DI" and txt_5_2.get() == "C"):
                produce_combine_p("Control_Group.xlsx", "Diabetes_Insulin_Group.xlsx", 1)
                messagebox.showinfo('Success!', res)
            elif (txt_5.get() == "D" and txt_5_2.get() == "DI") or (txt_5.get() == "DI" and txt_5_2.get() == "D"):
                produce_combine_p("Diabetes_Group.xlsx", "Diabetes_Insulin_Group.xlsx", 1)
                messagebox.showinfo('Success!', res)
            else:
                messagebox.showerror("Error", failed)

    btn_5 = Button(tab5, text="Generate", command=pval)
    btn_5.pack(padx=5, pady=5)

    # tab 6
    lbl_6 = Label(tab6, text="Group Names")
    lbl_6.pack(padx=2, pady=2)

    # b = IntVar()
    # radio_cd = Radiobutton(tab6, text="Control - Diabetes", variable=b, value=0)
    # radio_cd.pack(anchor=W)
    # radio_cdi = Radiobutton(tab6, text="Control - Diabetes+Insulin", variable=b, value=1)
    # radio_cdi.pack(anchor=W)
    # radio_ddi = Radiobutton(tab6, text="Diabetes - Diabetes+Insulin", variable=b, value=2)
    # radio_ddi.pack(anchor=W)

    txt_6 = Entry(tab6, width=40)
    txt_6.pack(padx=2, pady=2)

    txt_6_1 = Entry(tab6, width=40)
    txt_6_1.pack(padx=2, pady=2)

    a = IntVar()
    radio_up = Radiobutton(tab6, text="Up-regulated", variable=a, value=0)
    radio_up.pack(anchor=W)
    radio_down = Radiobutton(tab6, text="Down-regulated", variable=a, value=1)
    radio_down.pack(anchor=W)

    # generate up-regulated / down-regulated file based on radio choice
    def up_down():
        table = []
        res = "Yes! You have successfully generated the file"
        failed = "Please try again"
        if a.get() == 0:
            table = up_down_regulated(txt_6.get() + "_" + txt_6_1.get() + "_p_value_log_2FC.xlsx", 0)
            workbook = xlsxwriter.Workbook('Up (' + txt_6.get() + " x " + txt_6_1.get() + ').xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(table)):
                worksheet.write_column(0, x, table[x])
            workbook.close()
            messagebox.showinfo("Success!", res)
        elif a.get() == 1:
            table = up_down_regulated(txt_6.get() + "_" + txt_6_1.get() + "_p_value_log_2FC.xlsx", 1)
            workbook = xlsxwriter.Workbook('Down (' + txt_6.get() + " x " + txt_6_1.get() + ').xlsx')
            worksheet = workbook.add_worksheet()
            for x in range(len(table)):
                worksheet.write_column(0, x, table[x])
            workbook.close()
            messagebox.showinfo("Success!", res)
        else:
            messagebox.showerror("Failed!", failed)

    btn_6 = Button(tab6, text="Generate", command=up_down)
    btn_6.pack(padx=5, pady=5)

    # tab 7
    lbl_7 = Label(tab7, text="Group Name: Control, Diabetes, Diabetes+Insulin")
    lbl_7.pack(padx=5, pady=5)

    txt_7 = Entry(tab7, width=40)
    txt_7.pack(padx=2, pady=2)

    txt_7_1 = Entry(tab7, width=40)
    txt_7_1.pack(padx=2, pady=2)

    n = IntVar()
    radio_up = Radiobutton(tab7, text="Up-regulated", variable=n, value=0)
    radio_up.pack(anchor=W, padx=2)
    radio_down = Radiobutton(tab7, text="Down-regulated", variable=n, value=1)
    radio_down.pack(anchor=W, padx=2)

    lbl_7 = Label(tab7, text="Choose your adduct type(s)")
    lbl_7.place(x=600, y=40)

    adduct_types = StringVar()
    adduct_types.set(["M+H", "M+H-2H2O","M+H-H2O", "M+NH4", "M+Li", "M+NH4", "M+Na", "M+CH3OH+H", "M+K", "M+ACN+H", "M+2Na-H", ])

    lstbox = Listbox(tab7, listvariable=adduct_types, selectmode=MULTIPLE, width=20, height=10)
    lstbox.place(x=600, y=70)

    lbl_7 = Label(tab7, text="Type your tolerance number (in ppm): ")
    lbl_7.pack(side=LEFT)
    entry_7 = Entry(tab7, width=20)
    entry_7.pack(side=LEFT)

    # generate up-regulated / down-regulated file based on radio choice
    def automation():
        table = []
        res = "Yes! You have successfully generated the file"
        failed = "Please try again"
        selected = lstbox.curselection()

        lst = []
        for i in selected:
            lst.append(lstbox.get(i))

        if n.get() == 0:
            automate_db("Up (" + txt_7.get() + " x " + txt_7_1.get() + ").xlsx", lst, int(entry_7.get()), "HMDB_up(" + txt_7.get() + "x" + txt_7_1.get() + ")")
            messagebox.showinfo("Success!", res)

        elif n.get() == 1:
            automate_db("Down (" + txt_7.get() + " x " + txt_7_1.get() + ").xlsx", lst, int(entry_7.get()),
                        "HMDB_down(" + txt_7.get() + "x" + txt_7_1.get() + ")")
            messagebox.showinfo("Success!", res)
        else:
            messagebox.showerror("Failed!", failed)

    def csv_merged():
        table = []
        res = "Yes! You have successfully merged the files"
        failed = "Please try again"


        if n.get() == 0:
            if m.get() == 0:
                merge_csv("HMDB_up(CxDM1)", "HMDB_up(CxDM1).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 1:
                merge_csv("HMDB_up(CxDM1+I)", "HMDB_up(CxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 2:
                merge_csv("HMDB_up(DxDM1+I)", "HMDB_up(DxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            else:
                messagebox.showerror("Failed!", failed)
        elif n.get() == 1:
            if m.get() == 0:
                merge_csv("HMDB_down(CxDM1)", "HMDB_down(CxDM1).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 1:
                merge_csv("HMDB_down(CxDM1+I)", "HMDB_down(CxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 2:
                merge_csv("HMDB_down(DxDM1+I)", "HMDB_down(DxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            else:
                messagebox.showerror("Failed!", failed)
        else:
            messagebox.showerror("Failed!", failed)

    def kegg_automate():
        table = []
        res = "Yes! Automating process is successful."
        failed = "Please try again"

        if n.get() == 0:
            if m.get() == 0:
                automate_kegg_id("HMDB_up(CxDM1).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 1:
                automate_kegg_id("HMDB_up(CxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 2:
                automate_kegg_id("HMDB_up(DxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            else:
                messagebox.showerror("Failed!", failed)
        elif n.get() == 1:
            if m.get() == 0:
                automate_kegg_id("HMDB_down(CxDM1).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 1:
                automate_kegg_id("HMDB_down(CxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            elif m.get() == 2:
                automate_kegg_id("HMDB_down(DxDM1+I).csv")
                messagebox.showinfo("Success!", res)
            else:
                messagebox.showerror("Failed!", failed)
        else:
            messagebox.showerror("Failed!", failed)

    btn_7 = Button(tab7, text="Generate", command=automation)
    btn_7.place(x=330, y=300)
    btn_7 = Button(tab7, text="Merge files", command=csv_merged)
    btn_7.place(x=400, y=300)
    btn_7 = Button(tab7, text="Automate", command=kegg_automate)
    btn_7.place(x=480, y=300)
    window.mainloop()


# abbreviation
def abbreviation(table, string):
    if string == "Control":
       table.append("C")
    elif string == "Diabetes":
        table.append("DM1")
    elif string == "Diabetes+Insulin":
        table.append("DM1+I")
    elif string == "COF":
        table.append("COF")
    elif string == "PP":
        table.append("PP")
    elif string == "TAMM":
        table.append("TAMM")
    elif string == "PPF":
        table.append("PPF")
    elif string == "TAMF":
        table.append("TAMF")
    elif string == "QC":
        table.append("QC")
    elif string == "Before":
        table.append("Before")
    elif string == "After":
        table.append("After")
    elif string == "KSHV+HIV+":
        table.append("KSHV+HIV+")
    elif string == "KSHV+HIV-":
        table.append("KSHV+HIV-")
    elif string == "KSHV-HIV-":
        table.append("KSHV-HIV-")


# separate group
def separating_group(table, string, filename):
    count = 0
    tab = []
    original_table = read_data(filename)[0]
    slicing = slice(1, len(original_table))
    tab.append(original_table[slicing])

    for y in range(len(table)):
        record = []
        if table[y][0] == string:
            abbreviation(record, string)
            for x in range(len(table[1])):
                if isinstance(table[y][x], float) or table[y][x] is None:
                    record.append(table[y][x])
            new_record = record
            tab.append(new_record)

    count_table = ["Count"]
    for x in range(1, len(tab[0])): # not counting first row - title row
        count = 0
        for y in range(1, len(tab)): # not counting first column - m/z column
            if tab[y][x] is not None:
                count += 1
        count_table.append(count)

    print(tab)
    # appending the count table to count the appearance of data each row
    tab.append(count_table)
    print(count_table)
    return tab

# separate group for data_processed files
def data_separated_group(string, filename, new_filename):
    table = read_all_data(filename)
    tab = []
    tab.append(table[0])
    to_float = 0.0
    for y in range(len(table)):
        record = []
        if table[y][0] == string:
            abbreviation(record, string)
            for x in range(1, len(table[1])):
                # if isinstance(table[y][x], float) or table[y][x] is None:
                to_float = float(table[y][x])
                record.append(to_float)
            new_record = record
            tab.append(new_record)

    workbook = xlsxwriter.Workbook(new_filename)
    worksheet = workbook.add_worksheet()
    for x in range(len(tab)):
        worksheet.write_column(0, x, tab[x])
    workbook.close()



def final_separated_table(table):
    for x in range(len(table[0])):
        if table[5][x]/5.0 < 0.65:
            table[5][x] = 0
            for y in range(0, len(table) - 1):
                table[y][x] = None

    return table


# check if over 65%, if yes -> keep. If not, empty cells in row
def check_percentage(string):
    wb = openpyxl.load_workbook(filename=string)
    sheet = wb['Sheet1']
    row = sheet.max_row
    column = sheet.max_column

    print(column)
    if string == "COF_Muscle" or string == "PP_Muscle" or string == "TAMM_Muscle":
        for x in range(2, row + 1):
            if sheet.cell(row=x, column=8).value/6.0 < 0.65:
                sheet.cell(row=x, column=8).value = 0
                for y in range(2, column):
                    sheet.cell(row=x, column=y).value = None
    else:
        for x in range(2, row + 1):
            if sheet.cell(row=x, column=column).value/float(column - 2.0) < 0.65:
                sheet.cell(row=x, column=column).value = 0
                for y in range(2, column):
                    sheet.cell(row=x, column=y).value = None

    wb.save(string)


# remove rows if avg goes to 1
def remove_rows(table, i):
    # while i < len(table[0]):
    if table[len(table) - 1][i] == 1:
        for x in table:
            del x[i]
    # else:
    #     i += 1



# save to csv file
def save_csv(table, string):
    if string == "Raw":
        export_data = zip_longest(*table, fillvalue='')
        with open('final_data_raw.csv', 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)
            writer.writerows(export_data)
        file.close()
    elif string == "FL":
        export_data = zip_longest(*table, fillvalue='')
        with open('final_data_FL.csv', 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)
            writer.writerows(export_data)
        file.close()
    elif string == "FM":
        export_data = zip_longest(*table, fillvalue='')
        with open('final_data_FM.csv', 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)
            writer.writerows(export_data)
    elif string == "HIV":
        export_data = zip_longest(*table, fillvalue='')
        with open('final_data_HIV.csv', 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)
            writer.writerows(export_data)

        file.close()



# average of group data : Control_Group.xlsx, Diabetes_Group.xlsx, Diabetes_Insulin_Group.xlsx
def calculate_average(string):
    wb = openpyxl.load_workbook(filename=string)
    sheet = wb['Sheet1']
    row = sheet.max_row
    column = sheet.max_column
    col = []
    col.append("AVG")
    for x in range(2, row + 1):
        average = 0.0
        for y in range(2, column + 1):
            if sheet.cell(row=x, column=y).value is not None:
                average = average + sheet.cell(row=x, column=y).value
        print(average)
        average = average / float(column - 1.0)
        col.append(average)

    change_to_1(col)
    wb.save(string)

    return col



# change avg column to 1 if avg element = 0
def change_to_1(table):
    for x in range(len(table)):
        if table[x] == 0:
            table[x] = 1

    return table


# calculate log_2 of average of 2 group data
def calculate_log(str1, str2):
    avg_1 = calculate_average(str1)
    avg_2 = calculate_average(str2)

    log_col = []

    for x in range(1, len(avg_1)):
        log = get_log(avg_1[x], avg_2[x])
        log_col.append(log)

    return log_col


# change back empty cells to 0
def change_to_zero(table):
    # print(len(table))
    # print(len(table[0]))
    for x in range(len(table)):
        for y in range(len(table[0])):
            if table[x][y] == "":
                table[x][y] = 0


# test p value calculation
def test_p(data_1, data_2):
    # mean1 = np.mean(data_1)
    # mean2 = np.mean(data_2)
    # # get std error
    # se1 = sem(data_1)
    # se2 = sem(data_2)
    # # standard error on the difference between the samples
    # sed = np.sqrt(se1 ** 2.0 + se2 ** 2.0)
    # if sed == 0:
    #     return None
    # # calculate T Statistic
    # t_stat = (mean1 - mean2) / sed
    # # degrees of freedom
    # df = len(data_1) + len(data_2) - 2
    # # calculate the p-value
    # p = (1.0 - t.cdf(abs(t_stat), df))

    t_stat, p_val = stats.ttest_ind(data_1, data_2, equal_var=True)
    return p_val

    # var_a = data_1.var(ddof=1)
    # var_b = data_2.var(ddof=1)
    # N = len(data_1) + len(data_2)
    #
    # s = np.sqrt((var_a + var_b) / 2)
    # t = (data_1.mean() - data_2.mean()) / (s * np.sqrt(2 / N))
    # df = 2 * N - 2
    # p = 1 - stats.t.cdf(t, df=df)
    # return p


# get m/z column
def get_mz_col(string):
    return read_group_data(string)[0]


# p value to compare 2 groups
def get_p_value(str1, str2):
    # read data from specific groups
    table_1 = read_group_data(str1)
    slicing1 = slice(1, len(table_1))
    table_1 = table_1[slicing1]

    table_2 = read_group_data(str2)
    slicing2 = slice(1, len(read_group_data(str2)))
    table_2 = table_2[slicing2]

    # change None to 0 cell
    change_to_zero(table_1)
    change_to_zero(table_2)

    # p value calculation
    p_col = []

    # get each row
    data_1 = get_row(table_1)
    data_2 = get_row(table_2)

    # print(test_p(data_1[0], data_2[0]))
    for x in range(1, len(data_1)):
        p = test_p(data_1[x], data_2[x])
        p_col.append(p)

    return p_col


# test get log2
def get_log(a1, a2):
    return np.log2(a1) - np.log2(a2)


# combine p value with 2 group data
def produce_combine_p(str1, str2, type, filetype):
    # initialize table to combine data
    group_1 = ""
    group_2 = ""
    if str1 == "Control_Group.xlsx":
        group_1 = "C"
    elif str2 == "Control_Group.xlsx":
        group_2 = "C"
    if str1 == "Diabetes_Group.xlsx":
        group_1 = "DM1"
    elif str2 == "Diabetes_Group.xlsx":
        group_2 = "DM1"
    if str1 == "Diabetes_Insulin_Group.xlsx":
        group_1 = "DM1+I"
    elif str2 == "Diabetes_Insulin_Group.xlsx":
        group_2 = "DM1+I"
    if str1 == "After_Group.xlsx":
        group_1 = "After"
    elif str2 == "After_Group.xlsx":
        group_2 = "After"
    if str1 == "Before_Group.xlsx":
        group_1 = "Before"
    elif str2 == "Before_Group.xlsx":
        group_2 = "Before"
    if str1 == "COF_Group.xlsx" or "COF_new.xlsx":
        group_1 = "COF"
    elif str2 == "COF_Group.xlsx" or "COF_new.xlsx":
        group_2 = "COF"
    if str1 == "COF_Muscle_Group.xlsx":
        group_1 = "COF_Muscle"
    elif str2 == "COF_Muscle_Group.xlsx":
        group_2 = "COF_Muscle"
    if str1 == "PPF_Group.xlsx" or "PPF_new.xlsx":
        group_1 = "PPF"
    elif str2 == "PPF_Group.xlsx" or "PPF_new.xlsx":
        group_2 = "PPF"
    if str1 == "PP_Muscle_Group.xlsx":
        group_1 = "PP_Muscle"
    elif str2 == "PP_Muscle_Group.xlsx":
        group_2 = "PP_Muscle"
    if str1 == "KSHV+HIV+_Group.xlsx":
        group_1 = "KSHV+HIV+"
    elif str2 == "KSHV+HIV+_Group.xlsx":
        group_2 = "KSHV+HIV+"
    if str1 == "KSHV+HIV-_Group.xlsx":
        group_1 = "KSHV+HIV-"
    elif str2 == "KSHV+HIV-_Group.xlsx":
        group_2 = "KSHV+HIV-"
    if str1 == "KSHV-HIV-_Group.xlsx":
        group_1 = "KSHV-HIV-"
    elif str2 == "KSHV-HIV-_Group.xlsx":
        group_2 = "KSHV-HIV-"
    if str1 == "TAMF_new.xlsx":
        group_1 = "TAMF"
    elif str2 == "TAMF_new.xlsx":
        group_2 = "TAMF"



    table = []

    if filetype == "Raw":
        table.append(get_mz_col("Control_Group.xlsx"))
    elif filetype == "FL":
        table.append(get_mz_col("data_processed_FL_new.xlsx"))
    elif filetype == "FM":
        table.append(get_mz_col("data_processed_FM_new.xlsx"))
    elif filetype == "HIV":
        table.append(get_mz_col("data_processed_HIV_kNN_new.xlsx"))
    # read data from specific groups
    table_1 = read_group_data(str1)
    table_2 = read_group_data(str2)

    # append each column into the table
    # str1 data
    for x in range(1, len(table_1)):
        # insert name of group at the first row
        # table_1[x].insert(0, group_1)
        table.append(table_1[x])

    # str1 average data
    avg_1 = calculate_average(str1)
    change_to_1(avg_1)
    # avg_1.insert(0, "AVG")
    table.append(avg_1)

    # str2 data
    for x in range(1, len(table_2)):
        # table_2[x].insert(0, group_2)
        table.append(table_2[x])

    # str2 average data
    avg_2 = calculate_average(str2)
    change_to_1(avg_2)
    # avg_2.insert(0, "AVG")
    table.append(avg_2)

    # change None to 0 cell
    change_to_zero(table_1)
    change_to_zero(table_2)

    # append p_value column
    p_col = get_p_value(str1, str2)
    p_col.insert(0, "p_value")
    table.append(p_col)

    # New workbook xlsx file
    # p value only
    if type == 0:
        produce_file_p_log(table, str1, str2, type)

    # both p value and log
    elif type == 1:
        log_col = calculate_log(str1, str2)
        str_log = "LOG2FC " + group_1 + "/" + group_2
        log_col.insert(0, str_log)
        table.append(log_col)
        produce_file_p_log(table, str1, str2, type)


# decide if up-regulated or down-regulated
def up_down_regulated(file, up_or_down):
    table = read_all_data(file)
    change_to_zero(table)

    # keep track of index
    i = 1
    # if up-regulated
    if up_or_down == 0:
        # loop to put data to the correct list
        # for y in range(len(table[0]) - 1):

        while i < len(table[0]):
            if table[len(table) - 2][i] == 0 and table[len(table) - 1][i] == 0:
                for x in table:
                    del x[i]
            elif table[len(table) - 2][i] >= 0.05 or table[len(table) - 1][i] <= 0.5849:
                for x in table:
                    del x[i]
            else:
                i += 1

    # [[2,3,4,5], [3,4,5,6], [2,4,5,6]]
    # [[3,4,5], [4,5,6], [4,5,6]] - i = 1

    # if down-regulated
    elif up_or_down == 1:
        # loop to put data to the correct list
        while i < len(table[0]):
            if table[len(table) - 2][i] == 0 and table[len(table) - 1][i] == 0:
                for x in table:
                    del x[i]
            elif table[len(table) - 2][i] >= 0.05 or table[len(table) - 1][i] >= -0.5849:
                for x in table:
                    del x[i]
            else:
                i += 1
    return table




# remove first row then create new xlsx file
def remove_first_row(file_to_read, new_filename):
    table = read_all_data(file_to_read)
    for x in table:
        del x[0]

    workbook = xlsxwriter.Workbook(new_filename)
    worksheet = workbook.add_worksheet()
    for x in range(len(table)):
        worksheet.write_column(0, x, table[x])
    workbook.close()


# produce data in new file with p value or p value and log
def produce_file_p_log(table, str1, str2, type):
    if type == 0:
        workbook = xlsxwriter.Workbook(str1.replace("_new.xlsx", "") + "_" + str2.replace("_new.xlsx", "") + "_p_value.xlsx")
        worksheet = workbook.add_worksheet()
        for x in range(len(table)):
            worksheet.write_column(0, x, table[x])
        workbook.close()

    elif type == 1:
        workbook = xlsxwriter.Workbook(str1.replace("_new.xlsx", "") + "_" + str2.replace("_new.xlsx", "") + "_p_value_log_2FC.xlsx")
        worksheet = workbook.add_worksheet()
        for x in range(len(table)):
            worksheet.write_column(0, x, table[x])
        workbook.close()

# get std
def get_row(table):
    tab = []
    for x in range(len(table[0])):
        rec = []
        for y in range(len(table)):
            rec.append(table[y][x])
        new_rec = rec
        tab.append(new_rec)

    return tab

def csv_to_xlsx(string):
    df_new = pd.read_csv(string)
    writer = pd.ExcelWriter(string.replace('.csv', '.xlsx'))
    df_new.to_excel(writer, index=False)
    writer.save()



# automate on hmdb website
def automate_db(file, adduct, tolerance_number, file_name):
    # start = timeit.default_timer()
    table = read_all_data(file)
    # number of iterations for the automation
    i = 1
    j = 1
    x = 1
    # for x in range(1, len(table[0])):
    #     record = []
    #     while len(record) <= 700:
    #         record.append(table[x])
    #     hmdb.automate_hmdb(record)

    record = []
    while i <= math.ceil((len(table[0]) - 1) / 700):
        record = []
        j = 1
        while x < len(table[0]):
            if j > 700:
                break
            else:
                record.append(table[0][x])
                x += 1
                j += 1
        print(record)
        automate_hmdb(record, adduct, tolerance_number)
        time.sleep(3)
        os.rename("/Users/phucnguyen/Downloads/search.csv",
                  "/Users/phucnguyen/PycharmProjects/Metabolomic_Data/Automating_Excel/" + file_name + "_" + str(i) + ".csv")
        i += 1



    stop = timeit.default_timer()
    # print('Time: ', stop - start)


def browser_open(website_path):
    # add chrome driver to execute
    # To use this, you need to download chromedriver from https://chromedriver.chromium.org/downloads and choose
    # the version of google chrome you are using. Then, specify the path in executable variable like below.
    executable = {'executable_path': r'/Users/phucnguyen/Desktop/chromedriver'}

    options = webdriver.ChromeOptions()

    options.add_argument("--window-size=1400,900")
    options.add_argument("--start-maximized")

    options.add_argument("--disable-notification")

    browser = Browser('chrome', **executable, headless=False, options=options)

    browser.visit(website_path)

    return browser

# visit hmdb.ca to automate
def automate_hmdb(table, adduct, tolerance_number):
    # open hmdb.ca website
    browser = browser_open("https://hmdb.ca/spectra/ms/search")

    # find id for textarea - query_masses
    # query_mass = browser.find_by_id("query_masses")

    browser.fill("query_masses", '\n'.join(str(float(t) - 1.0) for t in table))
    adduct_type = browser.find_by_id("adduct_type")
    for a in adduct:
        adduct_type.select(a)

    browser.fill("tolerance", tolerance_number)

    tolerance = browser.find_by_id("tolerance_units")
    tolerance.select("ppm")

    # submit button -- search
    submit = browser.find_by_name("commit").first.click()
    # time.sleep(3)
    # download as csv
    submit_1 = browser.find_by_value("Download Results As CSV").first.click()


# automate by accessing to the map_pathway website
def automate_kegg_id(file, output_file):
    # get the kegg_id column from the hmdb excel file(s)
    storage = []
    visited = []
    with open(file, "r") as source:
        reader = csv.reader(source)
        with open(output_file, "w", newline='') as result:
            writer = csv.writer(result)
            for row in reader:
                if row[3] != "n/a" and row[3].replace(" ", "") not in storage:
                    storage.append(row[3])



    # access to pathway website to automate
    hm.automate_kegg(storage)



# merge csv files based on filename
def merge_csv(filename, output_file):
    path = "/Users/phucnguyen/PycharmProjects/Metabolomic_Data/Automating_Excel/"
    all_files = glob.glob(os.path.join(path, filename + "_*.csv"))
    df_merged = pd.concat([pd.read_csv(f) for f in all_files], ignore_index=True)
    df_merged.to_csv(output_file, index=False)






    # transpose.append(["Carbohydrate metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Lipid Metabolism", "Lipid Metabolism", "Lipid Metabolism", "Lipid Metabolism", "Lipid Metabolism", "Lipid Metabolism", "Metabolism of cofactors and vitamins", "Lipid Metabolism", "Energy Metabolism", "Amino acid metabolism", "Nucleotide metabolism", "Biosynthesis of other secondary metabolites", "Nucleotide metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism","Amino acid metabolism", "Metabolism of other amino acids", "Metabolism of other amino acids", "Metabolism of other amino acids", "Metabolism of other amino acids", "Metabolism of other amino acids", "Metabolism of other amino acids", "Metabolism of other amino acids",
    #                   "Carbohydrate metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Carbohydrate metabolism", "Biosynthesis of other secondary metabolites", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Lipid Metabolism", "Carbohydrate metabolism", "Glycan biosynthesis and metabolism", "Lipid Metabolism", "Lipid Metabolism", "Lipid Metabolism", "Lipid Metabolism",
    #                   "Lipid Metabolism", "Lipid Metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Glycan biosynthesis and metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Carbohydrate Metabolism", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of cofactors and vitamins", "Metabolism of terpenoids and polyketides", "Energy metabolism", "Energy metabolism", "Translation", "Xenobiotics biodegradation and metabolism", "Xenobiotics biodegradation and metabolism", "Xenobiotics biodegradation and metabolism", "Lipid Metabolism", "Global and overview maps", "Global and overview maps", "Global and overview maps", "Global and overview maps", "Global and overview maps", "Drug resistance: antineoplastic", "Drug resistance: antineoplastic", "Drug resistance: antineoplastic", "Drug resistance: antineoplastic", "Membrane Transport",
    #                   "Translation", "Translation", "Translation", "Translation", "Folding, sorting and degradation", "Transcription", "Transcription", "Replication and repair", "Transcription", "Folding, sorting and degradation",
    #                   "Folding, sorting and degradation", "Endocrine system", "Replication and repair", "Replication and repair", "Replication and repair", "Replication and repair", "Replication and repair", "Replication and repair", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Signaling molecules and interaction", "Signaling molecules and interaction", "Immune system", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Signaling molecules and interaction", "Cell growth and death", "Cell growth and death", "Cell growth and death", "Folding, sorting and degradation", "Folding, sorting and degradation", "Folding, sorting and degradation", "Transport and catabolism", "Transport and catabolism", "Transport and catabolism", "Folding, sorting and degradation", "Transport and catabolism", "Transport and catabolism", "Transport and catabolism", "Transport and catabolism", "Signal transduction", "Signal transduction", "Signal transduction", "Cell growth and death", "Aging", "Aging", "Cell growth and death", "Cell growth and death", "Cell growth and death", "Cell growth and death", "Circulatory system", "Circulatory system", "Circulatory system", "Signal transduction", "Signal transduction", "Signal transduction", "Signal transduction", "Development and regeneration",
    #                   "Signal transduction", "Signal transduction", "Development and regeneration", "Signal transduction", "Signal transduction", "Cellular community - eukaryotes", "Signaling molecules and interaction", "Signaling molecules and interaction", "Cellular community - eukaryotes", "Cellular community - eukaryotes", "Cellular community - eukaryotes", "Cellular community - eukaryotes", "Immune System", "Immune System", "Immune System", "Endocrine system", "Immune System", "Immune system", "Immune system", "Immune system", "Immune system", "Immune system", "Signal transduction", "Immune System", "Immune System", "Immune System", "Immune System", "Immune System", "Immune System", "Immune System", "Immune System", "Signal Transduction", "Immune System", "Immune System", "Immune System", "Environmental adaptation", "Environmental adaptation", "Environmental adaptation", "Nervous system", "Nervous system", "Nervous System", "Nervous System", "Nervous System", "Nervous System", "Nervous System", "Nervous System", "Nervous System", "Nervous System", "Sensory system", "Sensory system", "Sensory system", "Sensory system", "Cell motility", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine system", "Endocrine and metabolic disease", "Endocrine and metabolic disease", "Endocrine and metabolic disease", "Endocrine and metabolic disease", "Endocrine and metabolic disease", "Endocrine system", "Endocrine and metabolic disease",
    #                   "Endocrine and metabolic disease", "Excretory system", "Excretory system", "Excretory system", "Excretory system", "Excretory system", "Digestive system", "Digestive system", "Digestive system", "Digestive system", "Digestive system", "Digestive system", "Digestive system", "Digestive system", "Digestive system", "Digestive system", "Neurodegenerative disease", "Neurodegenerative disease", "Neurodegenerative disease", "Neurodegenerative disease", "Neurodegenerative disease", "Neurodegenerative disease", "Substance dependence", "Substance dependence", "Substance dependence", "Substance dependence", "Substance dependence", "Infectious disease: bacterial", "Infectious disease: bacterial", "Infectious disease: bacterial", "Infectious disease: bacterial", "Infectious disease: bacterial", "Infectious disease: parasitic", "Infectious disease: parasitic", "Infectious disease: parasitic", "Infectious disease: parasitic", "Infectious disease: parasitic", "Infectious disease: parasitic", "Infectious disease: bacterial", "Infectious disease: bacterial", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Infectious disease: viral", "Cancer: overview", "Cancer: overview", "Cancer: overview", "Cancer: overview", "Cancer: overview", "Cancer: overview", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types", "Cancer: specific types",
    #                   "Cancer: overview", "Cancer: overview", "Cancer: overview", "Immune disease", "Immune disease", "Immune disease", "Immune disease", "Immune disease", "Immune disease", "Immune disease", "Immune disease", "Cardiovascular disease", "Cardiovascular disease", "Cardiovascular disease", "Cardiovascular disease", "Cardiovascular disease"]) # stop at row 229, continue at row 230

    # workbook = xlsxwriter.Workbook('pathway_list.xlsx')
    # worksheet = workbook.add_worksheet()
    #
    # for x in range(3):
    #     worksheet.write_column(0, x, transpose[x])
    # workbook.close()


    # pathway = []

    # for x in range(len(list) - 1):
    #     pathway.append(list[x][1])
    #     print(pathway[x])


automate_kegg_id("HMDB_up(BeforexAfter)_1.csv", "new.csv")



# print(read_data("Raw_data_and_steps_Diabetes_data.xlsx")[0])

# print(separating_group(read_data("Raw_data_and_steps_Diabetes_data.xlsx"), "Control")[0])

# print(read_group_data_with_average("Control_Group.xlsx")[6])

# i = 0
# table = [[1,2,3,4, 3], [1,2,3,4, 3], [1,2,3,4, 3], [3,2,3,4,1]]
#
# print(remove_rows(table, i))

# r =requests.get('https://hmdb.ca/spectra/ms/search')
# query = {'query_masses': '123'}
# r = requests.post('https://hmdb.ca/spectra/ms/search', data= query)
#
# print(r.text)

# automate_db("Down (C x DM1).xlsx", ["M+H", "M+Li"], 10)

# os.rename("/Users/phucnguyen/PycharmProjects/Metabolomic_Data/Automating_Excel/search.csv", "/Users/phucnguyen/PycharmProjects/Metabolomic_Data/Automating_Excel/up.csv")

# tkinter_window()

# csv_to_xlsx('data_processed_HIV_kNN (no QCs).csv')

# remove_first_row("data_processed_HIV_kNN (no QCs).xlsx", "data_processed_HIV_kNN_new.xlsx")

# data_separated_group("COF", "data_processed_FL_new.xlsx", "COF_new.xlsx")

# print(calculate_average("COF_new.xlsx"))

# print(read_group_data("COF_new.xlsx")[5])

# produce_combine_p("COF_new.xlsx", "PPF_new.xlsx", 0, "FL")
# produce_combine_p("COF_new.xlsx", "TAMF_new.xlsx", 0, "FL")
# produce_combine_p("PPF_new.xlsx", "TAMF_new.xlsx", 0, "FL")
# produce_combine_p("COF_Muscle_new.xlsx", "PP_Muscle_new.xlsx", 0, "FM")
# produce_combine_p("COF_Muscle_new.xlsx", "TAMM_Muscle_new.xlsx", 0, "FM")
# produce_combine_p("PP_Muscle_new.xlsx", "TAMM_Muscle_new.xlsx", 0, "FM")
# produce_combine_p("Before_new.xlsx", "After_new.xlsx", 0, "HIV")
# produce_combine_p("Before_new.xlsx", "KSHV+HIV+_new.xlsx", 0, "HIV")
# produce_combine_p("Before_new.xlsx", "KSHV+HIV-_new.xlsx", 0, "HIV")
# produce_combine_p("Before_new.xlsx", "KSHV-HIV-_new.xlsx", 0, "HIV")
# produce_combine_p("After_new.xlsx", "KSHV+HIV+_new.xlsx", 0, "HIV")
# produce_combine_p("After_new.xlsx", "KSHV+HIV-_new.xlsx", 0, "HIV")
# produce_combine_p("After_new.xlsx", "KSHV-HIV-_new.xlsx", 0, "HIV")
# produce_combine_p("KSHV+HIV+_new.xlsx", "KSHV+HIV-_new.xlsx", 0, "HIV")
# produce_combine_p("KSHV+HIV+_new.xlsx", "KSHV-HIV-_new.xlsx", 0, "HIV")
# produce_combine_p("KSHV+HIV-_new.xlsx", "KSHV-HIV-_new.xlsx", 0, "HIV")

# produce_combine_p("COF_new.xlsx", "PPF_new.xlsx", 1, "FL")
# produce_combine_p("COF_new.xlsx", "TAMF_new.xlsx", 1, "FL")
# produce_combine_p("PPF_new.xlsx", "TAMF_new.xlsx", 1, "FL")
# produce_combine_p("COF_Muscle_new.xlsx", "PP_Muscle_new.xlsx", 1, "FM")
# produce_combine_p("COF_Muscle_new.xlsx", "TAMM_Muscle_new.xlsx", 1, "FM")
# produce_combine_p("PP_Muscle_new.xlsx", "TAMM_Muscle_new.xlsx", 1, "FM")
# produce_combine_p("Before_new.xlsx", "After_new.xlsx", 1, "HIV")
# produce_combine_p("Before_new.xlsx", "KSHV+HIV+_new.xlsx", 1, "HIV")
# produce_combine_p("Before_new.xlsx", "KSHV+HIV-_new.xlsx", 1, "HIV")
# produce_combine_p("Before_new.xlsx", "KSHV-HIV-_new.xlsx", 1, "HIV")
# produce_combine_p("After_new.xlsx", "KSHV+HIV+_new.xlsx", 1, "HIV")
# produce_combine_p("After_new.xlsx", "KSHV+HIV-_new.xlsx", 1, "HIV")
# produce_combine_p("After_new.xlsx", "KSHV-HIV-_new.xlsx", 1, "HIV")
# produce_combine_p("KSHV+HIV+_new.xlsx", "KSHV+HIV-_new.xlsx", 1, "HIV")
# produce_combine_p("KSHV+HIV+_new.xlsx", "KSHV-HIV-_new.xlsx", 1, "HIV")
# produce_combine_p("KSHV+HIV-_new.xlsx", "KSHV-HIV-_new.xlsx", 1, "HIV")


# Before - After
# Before - KSHV+HIV+
# Before - KSHV+HIV-
# Before - KSHV-HIV-
# After - KSHV+HIV+
# After - KSHV+HIV-
# After - KSHV-HIV-
# KSHV+HIV+ - KSHV+HIV-
# KSHV+HIV+ - KSHV-HIV-
# KSHV+HIV- - KSHV-HIV-